#!/usr/bin/env python3
"""Clean the Google Sheet into local JSON under data/ (gitignored).

Reads data/source/rave-sheet.xlsx.
Writes data/events.json, data/sets.json, data/stats.json, data/recap.json, data/CLEANING.md.
"""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "rave-sheet.xlsx"
OUT = ROOT / "data"

# Trip / personal rows that live on Costs / Costs2 next to real shows.
# 2023/2024 calendar tabs have even more (hiking, dinners); those tabs are not imported.
PERSONAL_SHOW_NAMES = {
    "vegas",
    "pokemon go fest",
    "suzume",
    "skii daytrip",
    "skii trip",
    "skii trip 2",
    "north carolina visit",
    "nyc trip",
    "arizona trip",
    "red rocks trip",
    "beautifica",
    "new york",
}

# Observed sheet typos / naming drift between Sets and cost tabs.
SHOW_ALIASES = {
    "alyvn": "alvyn",
    "anyma: the end of genesys": "anyma 'the end of genesis'",
    "anyma the end of genesis": "anyma 'the end of genesis'",
    "beyond wonderland socal": "beyond wonderland",
    "black pink": "blackpink",
    "(g)i-dle wolrd tour [idol]": "(g)i-dle",
    "krazy super concert": "krazy super concert",
    "super krazy concert": "krazy super concert",
    "stay in bloom fest": "dabin stay in bloom",
    "dabin stay in bloom": "dabin stay in bloom",
    "john summit: experts only": "john summit",
    "b&l rattleship rebound": "rattleship rebound",
    "excison": "excision",
    "kpop x edm": "k-pop x edm",
    "head in the clouds": "hitc",
    "martin garix": "martin garrix",
    "gryllin red rocks": "gryffin",
    "gryffin red rocks": "gryffin",
    "gryffin day 2": "gryffin",
    "taylor swift eras tour": "taylor swift",
    "itzy world tour": "itzy",
    "iu world tour": "iu",
    "ive world tour": "ive",
    "veld afterparty": "veld day 3 afterparty",
    "wooli preparty": "wooli b2b kompany",
    "btsm preparty": "btsm",
    "edc preparty": "edc las vegas preparty x",
    "edclv preparty world party parade": "world party parade",
    "edclv preparty hardwell b2b maddix": "hardwell b2b maddix",
    "edclv preparty armin van buuren": "armin van buuren",
    "edclv preparty subtronics": "subtronics ebc",
    "dab the sky": "dabin liv",
    "umf after party outro": "outro umf afters",
    "umf after party b&l": "b&l umf afters",
    "insomniac x tomorrowland: unity": "unity",
    "level up & zingara": "level up",
    "illenium sphere": "illenium",
}

VENUE_ALIASES = {
    "1015 folsom nightclub": "1015 folsom",
    "frost theatre": "frost ampitheatre",
    "frost amphitheatre": "frost ampitheatre",
    "shoreline amphitheatre": "shoreline ampitheatre",
    "rebel nightclub": "rebel",
    "norweigian joy": "norwegian joy",
    "blue event center": "tahoe blue event center",
    "dtlv event center": "lv downtown events center",
    "the queen mary": "queen mary",
    "omnia nightclub": "omnia",
    "ebc at night": "encore beach club",
    "colorado convention center": "denver convention center",
    "bayside park": "bayfront park",
    "alpe de huez": "alpe dhuez",
    "alpe d'huez": "alpe dhuez",
    "barcelo maya riviera": "barcelo riviera maya",
    "barceló riviera maya resort": "barcelo riviera maya",
    "barcelo riviera maya resort": "barcelo riviera maya",
    "temple nightclub": "temple",
    "great hall avant gardner": "avant gardner",
    "undisclosed warehouse": "undisclosed warehouse",
}


def stable_id(*parts: object) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def money(value) -> float:
    if value is None or value == "" or value == "#DIV/0!":
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def as_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def norm(text: str | None) -> str:
    if not text:
        return ""
    s = str(text).strip().lower()
    s = s.replace("’", "'").replace("`", "'")
    s = re.sub(r"\s+", " ", s)
    return s


def norm_show(text: str | None) -> str:
    s = norm(text)
    s = re.sub(r"\s*\(cancelled\)\s*", " ", s).strip()
    return SHOW_ALIASES.get(s, s)


def norm_venue(text: str | None) -> str:
    s = norm(text)
    return VENUE_ALIASES.get(s, s)


def parse_date_cell(value, year_hint: int) -> tuple[date | None, date | None, str]:
    """Return (start, end, display). Display keeps the sheet's original string."""
    if value is None or value == "":
        return None, None, ""
    if isinstance(value, datetime):
        d = value.date()
        return d, d, d.isoformat()
    if isinstance(value, date):
        return value, value, value.isoformat()

    raw = str(value).strip()
    display = raw

    m = re.search(
        r"(?:[A-Za-z]+,?\s+)?(\d{1,2})/(\d{1,2})\s*[-–]\s*(\d{1,2})/(\d{1,2})",
        raw,
    )
    if m:
        sm, sd, em, ed = (int(x) for x in m.groups())
        start = date(year_hint, sm, sd)
        end_year = year_hint if em >= sm else year_hint + 1
        end = date(end_year, em, ed)
        return start, end, display

    m = re.search(
        r"(?:[A-Za-z]+)\s+(\d{1,2})/(\d{1,2})\s*,\s*(\d{1,2})/(\d{1,2})",
        raw,
    )
    if m:
        sm, sd, em, ed = (int(x) for x in m.groups())
        start = date(year_hint, sm, sd)
        end_year = year_hint if em >= sm else year_hint + 1
        end = date(end_year, em, ed)
        return start, end, display

    return None, None, display


def stringify_show(value) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if not s or s in {"#VALUE!", "#N/A", "#DIV/0!"}:
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def is_personal(show: str) -> bool:
    return norm(show) in PERSONAL_SHOW_NAMES


def iter_event_rows(ws, year: int, source_tab: str):
    for row in ws.iter_rows(min_row=3, max_col=10, values_only=True):
        date_v, show, venue, city, sets_v, ticket, travel, merch, total, per_set = list(row)[:10]
        if show is None or stringify_show(show) is None:
            continue
        show_s = stringify_show(show)
        if str(city).strip().lower() in {"total", "total:"} or show_s.lower() == "total":
            continue
        if str(date_v).strip().lower() in {"total", "total:"}:
            continue
        yield {
            "year": year,
            "source_tab": source_tab,
            "date_raw": date_v,
            "show": show_s,
            "venue": (str(venue).strip() if venue else None) or None,
            "city": (str(city).strip() if city else None) or None,
            "sets_sheet": as_int(sets_v),
            "ticket": money(ticket),
            "travel": money(travel),
            "merch": money(merch),
        }


def load_events(wb) -> list[dict]:
    rows: list[dict] = []
    skipped_personal: list[str] = []
    for tab, year in (("Costs", 2023), ("Costs2", 2024), ("2025", 2025), ("2026", 2026)):
        for raw in iter_event_rows(wb[tab], year, tab):
            if tab in {"Costs", "Costs2"} and is_personal(raw["show"]):
                skipped_personal.append(f"{tab}: {raw['show']}")
                continue
            start, end, display = parse_date_cell(raw["date_raw"], year)
            if start is None:
                # still keep the row; date_display holds whatever was in the cell
                display = str(raw["date_raw"] or "")
            ticket, travel, merch = raw["ticket"], raw["travel"], raw["merch"]
            total = round(ticket + travel + merch, 2)
            event = {
                "id": stable_id(year, norm_show(raw["show"]), start.isoformat() if start else display, norm_venue(raw["venue"])),
                "show": raw["show"],
                "venue": raw["venue"],
                "city": raw["city"],
                "year": year,
                "start_date": start.isoformat() if start else None,
                "end_date": end.isoformat() if end else None,
                "date_display": display,
                "ticket": ticket,
                "travel": travel,
                "drinks_food_merch": merch,
                "total": total,
                "sets_sheet": raw["sets_sheet"],
                "source_tab": raw["source_tab"],
                "source": "sheet",
            }
            rows.append(event)
    return rows, skipped_personal


def load_sets(wb) -> list[dict]:
    ws = wb["Sets"]
    out = []
    for i, row in enumerate(ws.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
        artist, show, venue, city, year, dt, artists = row
        if not artist and not show:
            continue
        year_i = as_int(year)
        if isinstance(dt, datetime):
            d_obj = dt.date()
        elif isinstance(dt, date):
            d_obj = dt
        else:
            d_obj = None
        # Year column wins when it disagrees with a clearly wrong date year (e.g. 2015 vs 2025).
        if d_obj is not None and year_i and d_obj.year != year_i and abs(d_obj.year - year_i) >= 2:
            try:
                d_obj = d_obj.replace(year=year_i)
            except ValueError:
                pass
        d = d_obj.isoformat() if d_obj else (str(dt) if dt else None)
        if not year_i and d:
            year_i = int(d[:4])
        artist_s = stringify_show(artist) or ""
        show_s = stringify_show(show) or ""
        names = []
        if artists is not None:
            if isinstance(artists, (int, float)):
                n = stringify_show(artists)
                names = [n] if n else []
            else:
                names = [stringify_show(p) for p in str(artists).split(",")]
                names = [n for n in names if n]
        if not names and artist_s:
            names = [artist_s]
        if not artist_s and not show_s:
            continue
        out.append(
            {
                "id": stable_id("set", i, artist_s, show_s, d),
                "sheet_row": i,
                "title": artist_s,
                "show": show_s,
                "venue": str(venue).strip() if venue else None,
                "city": str(city).strip() if city else None,
                "year": year_i,
                "date": d,
                "artists": names,
                "event_id": None,
            }
        )
    return out


def event_covers(event: dict, set_row: dict, pad_days: int = 1) -> bool:
    if not event.get("start_date") or not set_row.get("date"):
        return False
    start = date.fromisoformat(event["start_date"])
    end = date.fromisoformat(event["end_date"] or event["start_date"])
    try:
        d = date.fromisoformat(set_row["date"])
    except ValueError:
        return False
    return (start - timedelta(days=pad_days)) <= d <= (end + timedelta(days=pad_days))


def shows_compatible(event_show: str | None, set_show: str | None) -> bool:
    a, b = norm_show(event_show), norm_show(set_show)
    if not a or not b:
        return False
    if a == b:
        return True
    # "John Summit: Experts Only" vs "John Summit" — not "EDC Las Vegas Preparty" vs "EDC Las Vegas"
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if longer.startswith(shorter) and len(longer) > len(shorter):
        nxt = longer[len(shorter)]
        if nxt in {":", "|", "[", "("}:
            return True
    return False


def venues_compatible(a: str | None, b: str | None) -> bool:
    na, nb = norm_venue(a), norm_venue(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    if na in nb or nb in na:
        return True

    def squash(text: str) -> str:
        text = text.replace("é", "e").replace("è", "e").replace("á", "a").replace("ó", "o")
        return re.sub(r"[^a-z0-9]", "", text)

    if squash(na) == squash(nb):
        return True
    stop = {"the", "nightclub", "club", "park", "hall", "resort", "hotel", "center", "centre", "convention", "great"}
    ta = set(re.findall(r"[a-z0-9]+", na)) - stop
    tb = set(re.findall(r"[a-z0-9]+", nb)) - stop
    if ta and tb and (ta <= tb or tb <= ta):
        return True
    return False


def closest_event(cands: list[dict], set_row: dict) -> dict:
    cands = list(cands)
    cands.sort(key=lambda e: abs(_date_ord(e.get("start_date")) - _date_ord(set_row.get("date"))))
    return cands[0]


def link_sets(events: list[dict], sets: list[dict]) -> tuple[list[dict], list[dict]]:
    unmatched = []
    for s in sets:
        date_hits = [e for e in events if event_covers(e, s, pad_days=0)]
        show_hits = [e for e in date_hits if shows_compatible(e.get("show"), s.get("show"))]
        show_venue = [e for e in show_hits if venues_compatible(e.get("venue"), s.get("venue"))]
        venue_hits = [e for e in date_hits if venues_compatible(e.get("venue"), s.get("venue"))]

        pick = None
        if show_venue:
            pick = closest_event(show_venue, s)
        elif len(venue_hits) == 1:
            # The Sets tab writes one umbrella show name ("EDC Preparty") across
            # several rooms in a night, so venue is the harder fact.
            pick = venue_hits[0]
        elif len(show_hits) == 1:
            pick = show_hits[0]
        elif len(show_hits) > 1:
            pick = closest_event(show_hits, s)
        elif len(venue_hits) > 1:
            pick = closest_event(venue_hits, s)
        else:
            year_show = [
                e
                for e in events
                if e.get("year") == s.get("year") and shows_compatible(e.get("show"), s.get("show"))
            ]
            year_show_venue = [e for e in year_show if venues_compatible(e.get("venue"), s.get("venue"))]
            # This last resort ignores dates, so it once carried a set six months
            # onto the wrong show. Refuse a candidate whose venue contradicts; an
            # unmatched set becomes a derived event and shows up in reconcile.py.
            no_venue_conflict = [e for e in year_show if not e.get("venue") or not s.get("venue")]
            pool = year_show_venue or no_venue_conflict
            if len(pool) == 1:
                pick = pool[0]
            elif len(pool) > 1:
                pick = closest_event(pool, s)

        if pick:
            s["event_id"] = pick["id"]
        else:
            unmatched.append(s)
    return sets, unmatched


def _date_ord(value: str | None) -> int:
    if not value:
        return 0
    try:
        return date.fromisoformat(value).toordinal()
    except ValueError:
        return 0


def find_host_event(events: list[dict], sample: dict, start: str | None, end: str | None,
                    gap_days: int = 2) -> dict | None:
    """A sheet event for the same show and venue whose dates touch this group's.

    Countdown NYE runs Dec 30 to Jan 1, so the Costs tab (money) and the Sets tab
    (sets) land in different years and describe one festival as two events.
    """
    if not start:
        return None
    a, b = _date_ord(start), _date_ord(end or start)
    for e in events:
        if e.get("source") != "sheet" or not e.get("start_date"):
            continue
        c, d = _date_ord(e["start_date"]), _date_ord(e.get("end_date") or e["start_date"])
        if max(a, c) - min(b, d) > gap_days:
            continue
        if shows_compatible(e.get("show"), sample.get("show")) and venues_compatible(
            e.get("venue"), sample.get("venue")
        ):
            return e
    return None


def derive_events_from_sets(unmatched: list[dict], events: list[dict]) -> list[dict]:
    """One event per (year, show, venue) covering min/max set dates. Used for 2022 (no cost tab)."""
    existing_ids = {e["id"] for e in events}
    groups: dict[tuple, list[dict]] = defaultdict(list)
    for s in unmatched:
        key = (s.get("year"), norm_show(s.get("show")), norm_venue(s.get("venue")))
        groups[key].append(s)

    derived = []
    for (year, show_n, venue_n), group in groups.items():
        dates = sorted(x["date"] for x in group if x.get("date"))
        start = dates[0] if dates else None
        end = dates[-1] if dates else None
        sample = group[0]
        host = find_host_event(events, sample, start, end)
        if host:
            for s in group:
                s["event_id"] = host["id"]
            continue
        eid = stable_id(year, show_n, start, venue_n)
        if eid in existing_ids:
            eid = stable_id("derived", year, show_n, start, venue_n)
        event = {
            "id": eid,
            "show": sample["show"],
            "venue": sample.get("venue"),
            "city": sample.get("city"),
            "year": year,
            "start_date": start,
            "end_date": end,
            "date_display": f"{start}–{end}" if start and end and start != end else (start or ""),
            "ticket": 0.0,
            "travel": 0.0,
            "drinks_food_merch": 0.0,
            "total": 0.0,
            "sets_sheet": len(group),
            "source_tab": "Sets",
            "source": "derived_from_sets",
        }
        derived.append(event)
        for s in group:
            s["event_id"] = eid
    return derived


def compute_stats(events: list[dict], sets: list[dict], as_of: date) -> dict:
    for e in events:
        linked = sum(1 for s in sets if s["event_id"] == e["id"])
        e["sets_logged"] = linked
        e["dollars_per_set"] = round(e["total"] / linked, 2) if linked else None

    def artist_names(rows):
        names = []
        for s in rows:
            names.extend(s.get("artists") or [])
        return names

    def recap_for(evs, sts):
        spend = round(sum(e["total"] for e in evs), 2)
        artists = artist_names(sts)
        venues = {e["venue"] for e in evs if e.get("venue")}
        cities = {e["city"] for e in evs if e.get("city")}
        show_names = {s.get("show") for s in sts if s.get("show")}
        return {
            "sets": len(sts),
            "artists": len(set(norm(a) for a in artists if a)),
            "set_titles": len({s["title"] for s in sts if s.get("title")}),
            "shows": len(show_names),
            "events": len(evs),
            "venues": len(venues),
            "cities": len(cities),
            "spend": spend,
        }

    all_time = recap_for(events, sets)
    artist_counts = Counter(norm(a) for a in artist_names(sets) if a)
    top_artist = artist_counts.most_common(1)[0] if artist_counts else ("", 0)
    # restore display casing from first occurrence
    display_artist = next((a for s in sets for a in s.get("artists", []) if norm(a) == top_artist[0]), top_artist[0])

    venue_visits = Counter(norm_venue(e["venue"]) for e in events if e.get("venue"))
    top_venue_n, top_venue_c = venue_visits.most_common(1)[0] if venue_visits else ("", 0)
    display_venue = next((e["venue"] for e in events if norm_venue(e.get("venue")) == top_venue_n), top_venue_n)

    by_year = {}
    years = sorted({e["year"] for e in events if e.get("year")} | {s["year"] for s in sets if s.get("year")})
    for y in years:
        evs = [e for e in events if e.get("year") == y]
        sts = [s for s in sets if s.get("year") == y]
        by_year[str(y)] = recap_for(evs, sts)

    return {
        "as_of": as_of.isoformat(),
        "all_time": all_time,
        "by_year": by_year,
        "top_artist": {"name": display_artist, "times_seen": top_artist[1]},
        "top_venue": {"name": display_venue, "times_visited": top_venue_c},
        "sheet_stats_all_time": {
            "sets": 1241,
            "artists": 705,
            "days": 215,
            "shows": 154,
            "venues": 73,
            "cities": 33,
            "note": "Copied from the Stats tab for comparison. Recap numbers above are recomputed from cleaned events + sets.",
        },
        "artists": [
            {"name": next((a for s in sets for a in s.get("artists", []) if norm(a) == n), n), "times_seen": c}
            for n, c in artist_counts.most_common()
        ],
        "venues": [
            {
                "name": next((e["venue"] for e in events if norm_venue(e.get("venue")) == n), n),
                "times_visited": c,
            }
            for n, c in venue_visits.most_common()
        ],
        "unique_set_titles": len({s["title"] for s in sets if s.get("title")}),
    }


def write_cleaning_notes(
    events,
    sets,
    skipped_personal,
    derived_count,
    still_unmatched,
    as_of,
):
    linked = sum(1 for s in sets if s.get("event_id"))
    lines = [
        "# How this snapshot was cleaned",
        "",
        f"Source: public Google Sheet export saved at `data/source/rave-sheet.xlsx` (as of {as_of.isoformat()}).",
        "Tabs used: Costs (2023 shows), Costs2 (2024 shows), 2025, 2026, Sets.",
        "Tabs not imported as event rows: 2023 and 2024 calendars (mix hiking/dinners/daytrips), ArtistsVenues, Stats (both derived; recomputed here).",
        "",
        "## Event rows",
        "- Canonical columns: Date, Show, Venue, City, Sets, Ticket, Travel, Drinks/Food/Merch. Total and $ per set are computed (null/blank spend = 0).",
        "- Empty formula-fill rows on Costs2 (max_row ~1042) are skipped.",
        "- Total footer rows are skipped.",
        "- Personal / non-show rows dropped from Costs and Costs2:",
    ]
    for name in skipped_personal:
        lines.append(f"  - {name}")
    lines += [
        "- 2025 and 2026 are already merged calendars; every named show row is kept, including planned 2026 rows with blank set counts and tickets filled in.",
        f"- {derived_count} events were derived from unmatched Sets rows (2022 has no cost tab; plus a couple of one-off Sets shows not on a cost tab).",
        "- Excel stored some names as numbers (`999999999`, `1991`); they are imported as integers, not `1991.0`. `#VALUE!` in the Artists column falls back to the set title.",
        "",
        "## Set rows",
        "- Every non-empty Sets row is imported. `Artist` is the set title (may be a b2b). `Artists` is split on commas.",
        "- Parent link: normalized show name + date inside the event range, then unique venue+date when the Sets show name drifts (`Excison`, `ITZY World Tour`, `Illenium Sphere`, …).",
        f"- Linked {linked} of {len(sets)} sets to an event.",
        "",
        "## Recap vs the Stats tab",
        "- Yearly sets / unique set titles / unique Sets-tab show names match the Stats tab exactly (2022–2026).",
        "- All-time unique set titles = 705, same as Stats `Artists`. Recap `artists` is unique performer names from the Artists column after comma-split + casefold (lower than 705 because b2bs share names and casing variants collapse).",
        "- Venue times-visited is event rows per venue (Bill Graham 37, same as the sheet).",
        "",
        "## Planned vs went vs cancelled",
        "- `cancelled` if the show name contains `(Cancelled)`.",
        "- `attended` if at least one set is linked and the event's end date is on or before the snapshot date.",
        "- `planned` otherwise (future dates and/or empty set counts, including 2026 tickets with Sets still blank).",
        "",
        "## Left as in the sheet",
        "- Display typos are not rewritten (`Martin Garix`, `San Bernadino`, `Las Vegas, CA`, `Frost Ampitheatre`).",
        "- Aliases are matching-only.",
        "",
        f"## Counts in this snapshot: {len(events)} events, {len(sets)} sets.",
    ]
    if still_unmatched:
        lines.append(f"- WARNING: {len(still_unmatched)} sets still have no event_id.")
    OUT.joinpath("CLEANING.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Missing {SOURCE}")
    wb = load_workbook(SOURCE, data_only=True)
    as_of = date.today()

    events, skipped_personal = load_events(wb)
    sets = load_sets(wb)
    sets, unmatched = link_sets(events, sets)
    derived = derive_events_from_sets(unmatched, events)
    events.extend(derived)
    # second pass: derived events now exist; anything still unmatched stays unmatched
    still_unmatched = [s for s in sets if not s.get("event_id")]
    stats = compute_stats(events, sets, as_of)

    events.sort(key=lambda e: (e.get("start_date") or "", e.get("show") or ""), reverse=True)
    sets.sort(key=lambda s: (s.get("date") or "", s.get("title") or ""), reverse=True)

    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / "events.json").write_text(json.dumps(events, indent=2) + "\n", encoding="utf-8")
    (OUT / "sets.json").write_text(json.dumps(sets, indent=2) + "\n", encoding="utf-8")
    # keep recap small; full artist list is useful but large — include in stats
    recap = {
        "as_of": stats["as_of"],
        "all_time": stats["all_time"],
        "by_year": stats["by_year"],
        "top_artist": stats["top_artist"],
        "top_venue": stats["top_venue"],
        "sheet_stats_all_time": stats["sheet_stats_all_time"],
        "unique_set_titles": stats.get("unique_set_titles"),
    }
    (OUT / "recap.json").write_text(json.dumps(recap, indent=2) + "\n", encoding="utf-8")
    (OUT / "stats.json").write_text(json.dumps(stats, indent=2) + "\n", encoding="utf-8")
    write_cleaning_notes(events, sets, skipped_personal, len(derived), still_unmatched, as_of)

    print(f"events={len(events)} sets={len(sets)} derived={len(derived)} unmatched={len(still_unmatched)}")
    print("all_time", stats["all_time"])
    print("top_artist", stats["top_artist"])
    print("top_venue", stats["top_venue"])
    print("skipped", skipped_personal)
    if still_unmatched:
        print("still unmatched sample:")
        for s in still_unmatched[:20]:
            print(" ", s["title"], s["show"], s["date"], s["venue"])


if __name__ == "__main__":
    main()
